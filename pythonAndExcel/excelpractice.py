import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.sheetnames)
sheet = wb['Sheet3']
print(sheet.title)

actvsheet = wb.active
print(actvsheet)

# Getting Cells from the Sheets
print('\n')

sheet = wb['Sheet1']
print(type(sheet))

for i in sheet.__iter__():
    for j in i.__iter__():
        print(j.value)
    print('\n')


print('\n')

print(sheet['A1'])

print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print(f'Row {c.row}, Column {c.column}, is {c.value} ')

# row and column can be accessed with number starting frm 1
print('\n')

print(sheet.cell(row=1, column=1))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=1, column=2).value)

print('\n')

print(sheet.max_row)
print(sheet.max_column)

print('\n')

# Converting Between Column Letters and Numbers
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))
print(column_index_from_string('AA'))

print('\n')
print(tuple(sheet['A1':'C3']))
print('\n')

for rowOfCellObj in sheet['A1':'C3']:
    for cellObj in rowOfCellObj:
        print(cellObj.coordinate, cellObj.value)
    print('---end of row---')

 # create and save spreadsheet doc
newWb = openpyxl.Workbook()
newWb.sheetnames
s = newWb.active
print(f'initial name: {s}')
s.title = 'eggs and bacon'
print(newWb.sheetnames)
newWb.save('new.xlsx')

# creating and removing sheets
newWb.create_sheet()
newWb.create_sheet()
newWb.create_sheet(index=0, title='first sheet at idx 0')
newWb.create_sheet(index=0, title='second at idx 0')
newWb.create_sheet(index=3, title='d')

del newWb['d']

# newWb.save('new.xlsx')


print(newWb.sheetnames)

print('\n')
# writing values in cells
wbb = openpyxl.Workbook()
sheet = wbb['Sheet']
print(sheet)
sheet['A1'] = 'hello, world'
print(sheet['A1'].value)

# setting font of a cell
from openpyxl.styles import Font

italic24Font = Font(size=24, italic=True, bold=True, name='Times New Roman') # font created 
sheet['A1'].font = italic24Font
sheet['A1'] = 'Hello World1'
wbb.save('styles.xlsx')


print('\n')
# using formulas
wbbb = openpyxl.Workbook()
sheet = wbbb.active
sheet['A1'] = 200
sheet['A2'] = 300

sheet['A3'] = '=SUM(A1:A2)'
wb.save('formula.xlsx')


# Setting Row Height and Column Width
print('\n')
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'

sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].Width = 20
wbbb.save('dimentions.xlsx')

# Merging and Unmerging Cells
print(wb.active)
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'two merged cell'
wbbb.save('merged.xlsx')

#sheet.unmerge_cells('C5:D5')


# Freezing Panes
wbCurr = openpyxl.load_workbook('produceSales.xlsx')
sheetCurr = wbCurr.active
sheet.freeze_panes = 'A2' # freezing the rows above A2
wbCurr.save('freeze.xlsx')


# charts

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')

chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)

sheet.add_chart(chartObj, 'C5')
wb.save('chart.xlsx')